"""Main application window — assembles all panels, handles analysis logic."""

from __future__ import annotations

import traceback
import math
from pathlib import Path

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QPushButton,
    QTabWidget, QStatusBar, QMenuBar, QMessageBox, QFileDialog,
    QSplitter, QLabel, QComboBox, QApplication, QProgressDialog, QScrollArea,
    QInputDialog,
)
from PySide6.QtCore import Qt, QThread, Signal, QObject, QSize
from PySide6.QtGui import QAction

import luma
from luma.i18n.translator import t, init as i18n_init, set_language, AVAILABLE_LANGUAGES, get_language
from luma.gui.widgets.coord_input import LatitudeInput, LongitudeInput, RadiusInput
from luma.gui.widgets.source_selector import SourceSelector
from luma.gui.widgets.map_viewer import MapViewer
from luma.gui.widgets.results_table import (
    ResultsTable, MetricsPanel, SummaryPanel, WarningsPanel,
)
from luma.gui.widgets.help_bubble import HelpBubble, labeled_input_with_help
from luma.gui.widgets.temporal_panel import TemporalPanel
from luma.gui.widgets.compare_panel import ComparePanel
from luma.gui.widgets.aoi_widget import AOIWidget
from luma.gui.widgets.objective_selector import ObjectiveSelector
from luma.gui.widgets.aoi_comparison_panel import AOIComparisonPanel
from luma.gui.task_runner import TaskWorker
from luma.gui.dialogs.about import AboutDialog
from luma.gui.dialogs.settings import SettingsDialog
from luma.core.raster import (
    clip_raster_to_buffer, clip_raster_to_geometry,
    align_raster_pair, align_raster_to_reference,
)
from luma.core.aoi import AOI
from luma.core.stats import (
    compute_class_statistics, compute_landscape_metrics,
    generate_quality_warnings, AnalysisResult, compute_transition_matrix,
    validate_legend_classes,
)
from luma.core.buffer import buffer_geojson, buffer_area_km2
from luma.core.project import build_project, load_project, save_project
from luma.output.serialization import serializable_parameters
from luma.output.charts import (
    build_compare_gradient_figure, build_temporal_series_figure, save_figure,
)
from luma.output.geospatial import export_points_buffers
from luma.sources.catalog import (
    load_legend_classes, get_source, load_legend, resolve_remote_url, validate_year,
)


# ---------------------------------------------------------------------------
# Worker for background analysis
# ---------------------------------------------------------------------------
class AnalysisWorker(QObject):
    finished = Signal(object)  # AnalysisResult or Exception
    progress = Signal(str)

    def __init__(self, source_path: str, lon: float, lat: float,
                 radius_m: float, legend_key: str, aoi: AOI | None = None):
        super().__init__()
        self.source_path = source_path
        self.lon = lon
        self.lat = lat
        self.radius_m = radius_m
        self.legend_key = legend_key
        self.aoi = aoi

    def run(self) -> None:
        try:
            self.progress.emit(t("status.analyzing"))
            legend_classes = load_legend_classes(self.legend_key)
            legend_meta = load_legend(self.legend_key)

            if self.aoi is not None:
                raster = clip_raster_to_geometry(self.source_path, self.aoi)
            else:
                raster = clip_raster_to_buffer(
                    self.source_path, self.lon, self.lat, self.radius_m
                )
            valid_mask = raster.valid_mask_for_legend(legend_classes)
            legend_validation = validate_legend_classes(
                raster.data, valid_mask, legend_classes
            )

            class_stats = compute_class_statistics(
                raster.data, valid_mask, raster.pixel_area_m2, legend_classes
            )
            landscape = compute_landscape_metrics(
                class_stats, raster.data, valid_mask, raster.pixel_area_m2
            )
            warning_radius = self.radius_m
            if self.aoi is not None:
                warning_radius = math.sqrt(self.aoi.area_m2 / math.pi)
            warnings = generate_quality_warnings(
                int(valid_mask.sum()), raster.pixel_area_m2, warning_radius
            )
            valid_pixels = int(valid_mask.sum())

            result = AnalysisResult(
                class_stats=class_stats,
                landscape_metrics=landscape,
                total_area_m2=valid_pixels * raster.pixel_area_m2,
                total_valid_pixels=valid_pixels,
                pixel_area_m2=raster.pixel_area_m2,
                quality_warnings=warnings,
                source_name=legend_meta.get("name", self.legend_key),
                source_accuracy=legend_meta.get("reported_accuracy", "N/A"),
                raster_data=raster.data,
                raster_valid_mask=valid_mask,
                raster_transform=raster.transform,
                raster_crs=raster.crs,
                provenance={
                    "legend_key": self.legend_key,
                    "crs": str(raster.crs),
                    "pixel_area_m2": raster.pixel_area_m2,
                    "width": int(raster.data.shape[1]),
                    "height": int(raster.data.shape[0]),
                    "unknown_class_ids": legend_validation.unknown_ids,
                },
            )
            self.finished.emit(result)

        except Exception as exc:
            self.finished.emit(exc)


class ResponsiveTabWidget(QTabWidget):
    """Tab widget whose labels never dictate the application minimum width."""

    def minimumSizeHint(self) -> QSize:  # noqa: N802 - Qt API
        return QSize(360, 280)


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------
class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        # Portuguese is the default because the primary audience is Brazilian
        # academic users; English remains available through Settings.
        i18n_init("pt_BR")
        self._last_result: AnalysisResult | None = None
        self._last_params: dict = {}
        self._last_temporal: dict | None = None
        self._last_temporal_years: tuple[int, int] = (0, 0)
        self._last_temporal_series: list[dict] | None = None
        self._last_compare: list[dict] | None = None
        self._last_compare_points: list[dict] | None = None
        self._aoi: AOI | None = None
        self._active_task_thread: QThread | None = None
        self._active_task_worker: TaskWorker | None = None
        self._auto_compact = False
        self._setup_ui()
        self._setup_menu()
        self.setMinimumSize(520, 420)
        self.setWindowTitle(t("app.title"))
        self._status_bar = QStatusBar()
        self.setStatusBar(self._status_bar)
        self._status_bar.showMessage(t("status.ready"))

    # ── UI setup ──────────────────────────────────────────────────────────

    def _setup_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        self._main_splitter = splitter
        root.addWidget(splitter)

        # ── Left panel (input) ────────────────────────────────────────────
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(8, 8, 8, 8)

        self._objective_selector = ObjectiveSelector()
        self._objective_selector.objective_changed.connect(self._on_objective_changed)
        left_layout.addWidget(self._objective_selector)

        # Coordinate inputs
        self._lat_input = LatitudeInput(-23.55)
        self._lon_input = LongitudeInput(-46.63)
        self._radius_input = RadiusInput(5000)

        lay, self._lbl_lat, self._tip_lat = labeled_input_with_help(
            t("input.latitude"), self._lat_input, t("tips.latitude")
        )
        left_layout.addLayout(lay)
        lay, self._lbl_lon, self._tip_lon = labeled_input_with_help(
            t("input.longitude"), self._lon_input, t("tips.longitude")
        )
        left_layout.addLayout(lay)
        lay, self._lbl_rad, self._tip_rad = labeled_input_with_help(
            t("input.radius"), self._radius_input, t("tips.radius")
        )
        left_layout.addLayout(lay)

        self._aoi_widget = AOIWidget()
        self._aoi_widget.canvas.setMinimumSize(260, 130)
        self._aoi_widget.setMaximumHeight(245)
        self._aoi_widget.set_extent((-46.8, -23.7, -46.5, -23.4))
        self._aoi_widget.aoi_changed.connect(self._on_aoi_changed)
        left_layout.addWidget(self._aoi_widget)

        # Source selector
        self._source_selector = SourceSelector()
        left_layout.addWidget(self._source_selector)

        # Analyze button
        self._btn_analyze = QPushButton(t("input.analyze"))
        self._btn_analyze.setStyleSheet(
            "QPushButton { background: #1b7f3b; color: white; padding: 10px; "
            "border-radius: 5px; font-size: 15px; font-weight: bold; }"
            "QPushButton:hover { background: #176b32; }"
            "QPushButton:focus { border: 2px solid #0b3d1b; }"
            "QPushButton:disabled { background: #95a5a6; }"
        )
        self._btn_analyze.clicked.connect(self._run_analysis)

        left_layout.addStretch()
        left.setMinimumWidth(260)
        left.setMaximumWidth(370)
        self._input_form = left
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        left_scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        left_scroll.setWidget(left)
        self._left_column = QWidget()
        left_column_layout = QVBoxLayout(self._left_column)
        left_column_layout.setContentsMargins(0, 0, 0, 8)
        left_column_layout.setSpacing(6)
        left_column_layout.addWidget(left_scroll, stretch=1)
        self._btn_compact_results = QPushButton(t("input.view_results"))
        self._btn_compact_results.clicked.connect(
            lambda: self._set_input_panel_visible(False)
        )
        self._btn_compact_results.setVisible(False)
        left_column_layout.addWidget(self._btn_compact_results)
        left_column_layout.addWidget(self._btn_analyze)
        self._left_column.setMinimumWidth(260)
        self._left_column.setMaximumWidth(370)
        splitter.addWidget(self._left_column)

        # ── Right panel (tabs: results / temporal / compare) ──────────────
        right = QWidget()
        self._right_panel = right
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(4, 4, 4, 4)

        self._compact_controls = QWidget()
        compact_layout = QHBoxLayout(self._compact_controls)
        compact_layout.setContentsMargins(0, 0, 0, 0)
        compact_layout.addStretch()
        self._btn_open_inputs = QPushButton(t("input.parameters"))
        self._btn_open_inputs.setStyleSheet(
            "QPushButton { background: #176ca6; color: white; padding: 7px 14px; "
            "border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background: #125781; }"
        )
        self._btn_open_inputs.clicked.connect(
            lambda: self._set_input_panel_visible(True)
        )
        compact_layout.addWidget(self._btn_open_inputs)
        self._compact_controls.setVisible(False)
        right_layout.addWidget(self._compact_controls)

        self._tabs = ResponsiveTabWidget()
        self._tabs.setDocumentMode(True)
        self._tabs.setElideMode(Qt.TextElideMode.ElideRight)
        self._tabs.tabBar().setUsesScrollButtons(True)
        self._tabs.currentChanged.connect(self._on_tab_changed)

        # -- Tab 1: Single Analysis --
        tab_single = QWidget()
        tab_single_layout = QVBoxLayout(tab_single)
        tab_single_layout.setContentsMargins(0, 0, 0, 0)

        # Map + results split vertically
        inner_splitter = QSplitter(Qt.Orientation.Vertical)

        self._map_viewer = MapViewer()
        self._map_viewer.setMinimumHeight(220)
        inner_splitter.addWidget(self._map_viewer)

        results_widget = QWidget()
        results_layout = QVBoxLayout(results_widget)
        results_layout.setContentsMargins(0, 0, 0, 0)

        self._warnings_panel = WarningsPanel()
        results_layout.addWidget(self._warnings_panel)

        self._summary_panel = SummaryPanel()
        results_layout.addWidget(self._summary_panel)

        self._results_table = ResultsTable()
        self._metrics_panel = MetricsPanel()

        self._result_tabs = QTabWidget()
        self._result_tabs.setDocumentMode(True)
        self._result_tabs.addTab(self._results_table, t("tabs.coverage"))
        metrics_scroll = QScrollArea()
        metrics_scroll.setWidgetResizable(True)
        metrics_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        metrics_scroll.setWidget(self._metrics_panel)
        self._result_tabs.addTab(metrics_scroll, t("tabs.metrics"))
        results_layout.addWidget(self._result_tabs)

        inner_splitter.addWidget(results_widget)
        inner_splitter.setChildrenCollapsible(False)
        inner_splitter.setStretchFactor(0, 3)
        inner_splitter.setStretchFactor(1, 2)
        inner_splitter.setSizes([420, 300])

        tab_single_layout.addWidget(inner_splitter)
        single_scroll = QScrollArea()
        single_scroll.setWidgetResizable(True)
        single_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        single_scroll.setWidget(tab_single)
        self._tabs.addTab(single_scroll, t("tabs.single"))

        # -- Tab 2: Temporal Analysis --
        self._temporal_panel = TemporalPanel()
        self._temporal_panel.analyze_requested.connect(self._run_temporal_analysis)
        self._temporal_panel.analyze_multi_requested.connect(self._run_temporal_series)
        temporal_scroll = QScrollArea()
        self._temporal_scroll = temporal_scroll
        temporal_scroll.setWidgetResizable(True)
        temporal_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        temporal_scroll.setWidget(self._temporal_panel)
        self._tabs.addTab(temporal_scroll, t("tabs.temporal"))

        # -- Tab 3: Compare Points --
        self._compare_panel = ComparePanel()
        self._compare_panel.compare_requested.connect(self._run_comparison)
        self._compare_panel.open_map_requested.connect(self._on_compare_open_map)
        self._compare_panel.bulk_download_requested.connect(self._on_compare_bulk_download)
        self._compare_panel.map_tiff_requested.connect(self._on_compare_map_tiff)
        compare_scroll = QScrollArea()
        compare_scroll.setWidgetResizable(True)
        compare_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        compare_scroll.setWidget(self._compare_panel)
        self._tabs.addTab(compare_scroll, t("tabs.compare"))

        # -- Tab 4: Compare Map --
        self._compare_map_viewer = MapViewer()
        self._tabs.addTab(self._compare_map_viewer, t("tabs.compare_map"))

        # -- Tab 5: Compare polygonal AOIs --
        aoi_compare_tab = QWidget()
        aoi_compare_layout = QVBoxLayout(aoi_compare_tab)
        self._aoi_compare_panel = AOIComparisonPanel()
        self._aoi_compare_panel.analyze_requested.connect(self._run_aoi_comparison)
        self._aoi_compare_results = ResultsTable()
        aoi_compare_layout.addWidget(self._aoi_compare_panel, stretch=3)
        aoi_compare_layout.addWidget(self._aoi_compare_results, stretch=2)
        aoi_compare_scroll = QScrollArea()
        aoi_compare_scroll.setWidgetResizable(True)
        aoi_compare_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        aoi_compare_scroll.setWidget(aoi_compare_tab)
        self._tabs.addTab(aoi_compare_scroll, t("tabs.compare_aois"))

        right_layout.addWidget(self._tabs)
        splitter.addWidget(right)
        splitter.setChildrenCollapsible(False)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([320, 880])

    def _setup_menu(self) -> None:
        menu_bar = self.menuBar()
        menu_bar.clear()

        # File menu
        file_menu = menu_bar.addMenu(t("menu.file"))
        save_project_action = QAction(t("menu.save_project"), self)
        save_project_action.triggered.connect(self._save_project)
        file_menu.addAction(save_project_action)

        open_project_action = QAction(t("menu.open_project"), self)
        open_project_action.triggered.connect(self._open_project)
        file_menu.addAction(open_project_action)
        file_menu.addSeparator()

        export_csv = QAction(t("menu.export_csv"), self)
        export_csv.triggered.connect(self._export_csv)
        file_menu.addAction(export_csv)

        export_xlsx = QAction(t("menu.export_xlsx"), self)
        export_xlsx.triggered.connect(self._export_xlsx)
        file_menu.addAction(export_xlsx)

        export_json = QAction(t("menu.export_json"), self)
        export_json.triggered.connect(self._export_json)
        file_menu.addAction(export_json)

        export_pdf = QAction(t("menu.export_pdf"), self)
        export_pdf.triggered.connect(self._export_pdf)
        file_menu.addAction(export_pdf)

        export_tiff = QAction(t("menu.export_compare_tiff"), self)
        export_tiff.triggered.connect(self._export_compare_tiff)
        file_menu.addAction(export_tiff)

        export_geospatial = QAction(t("menu.export_geospatial"), self)
        export_geospatial.triggered.connect(self._export_geospatial_bundle)
        file_menu.addAction(export_geospatial)

        export_chart = QAction(t("menu.export_chart"), self)
        export_chart.triggered.connect(self._export_chart)
        file_menu.addAction(export_chart)

        file_menu.addSeparator()
        exit_act = QAction(t("menu.exit"), self)
        exit_act.triggered.connect(self.close)
        file_menu.addAction(exit_act)

        # Settings menu
        settings_menu = menu_bar.addMenu(t("menu.settings"))
        settings_act = QAction(t("menu.settings"), self)
        settings_act.triggered.connect(self._open_settings)
        settings_menu.addAction(settings_act)

        view_menu = menu_bar.addMenu(t("menu.view"))
        self._toggle_inputs_action = QAction(t("menu.input_panel"), self)
        self._toggle_inputs_action.setCheckable(True)
        self._toggle_inputs_action.setChecked(not self._left_column.isHidden())
        self._toggle_inputs_action.triggered.connect(self._set_input_panel_visible)
        view_menu.addAction(self._toggle_inputs_action)

        # Help menu
        help_menu = menu_bar.addMenu(t("menu.help"))
        about_act = QAction(t("menu.about"), self)
        about_act.triggered.connect(self._show_about)
        help_menu.addAction(about_act)

    # ── Analysis logic ────────────────────────────────────────────────────

    def _on_aoi_changed(self, aoi: AOI | None) -> None:
        self._aoi = aoi
        if aoi is not None:
            lat, lon = aoi.centroid_wgs84
            self._lat_input.setValue(lat)
            self._lon_input.setValue(lon)
            self._map_viewer.show_aoi(aoi, center=(lat, lon))

    def _on_objective_changed(self, objective: str) -> None:
        """Open the tab matching the user's research question."""
        tab_by_objective = {"single": 0, "temporal": 1, "compare": 2, "compare_aois": 4}
        index = tab_by_objective.get(objective)
        if index is not None:
            self._tabs.setCurrentIndex(index)

    def _on_tab_changed(self, index: int) -> None:
        objective_by_tab = {
            0: "single", 1: "temporal", 2: "compare", 4: "compare_aois",
        }
        objective = objective_by_tab.get(index)
        if objective:
            self._objective_selector.set_objective(objective)

    def _apply_input_panel_visibility(self, visible: bool) -> None:
        compact = self.width() < 720
        maximum_input_width = 16_777_215 if compact and visible else 370
        self._input_form.setMaximumWidth(maximum_input_width)
        self._left_column.setMaximumWidth(maximum_input_width)
        self._left_column.setVisible(visible)
        self._right_panel.setVisible(not visible if compact else True)
        self._compact_controls.setVisible(compact and not visible)
        self._btn_compact_results.setVisible(compact and visible)
        action = getattr(self, "_toggle_inputs_action", None)
        if action is not None:
            action.blockSignals(True)
            action.setChecked(visible)
            action.blockSignals(False)
        if visible:
            total = max(self._main_splitter.width(), 800)
            self._main_splitter.setSizes([min(320, total // 3), total])

    def _set_input_panel_visible(self, visible: bool) -> None:
        self._auto_compact = self.width() < 720
        self._apply_input_panel_visibility(visible)

    def resizeEvent(self, event) -> None:  # noqa: N802 - Qt API
        super().resizeEvent(event)
        if not hasattr(self, "_left_column"):
            return
        width = event.size().width()
        if width < 720 and not self._left_column.isHidden():
            self._auto_compact = True
            self._apply_input_panel_visibility(False)
        elif width >= 840 and self._auto_compact:
            self._auto_compact = False
            self._apply_input_panel_visibility(True)

    def _run_analysis(self) -> None:
        if self._active_task_thread is not None or (
            getattr(self, "_thread", None) is not None and self._thread.isRunning()
        ):
            QMessageBox.information(self, "Análise em andamento", "Aguarde a análise atual terminar.")
            return
        lat = self._lat_input.value()
        lon = self._lon_input.value()
        radius = self._radius_input.value()
        aoi = self._aoi
        legend_key = self._source_selector.get_legend_key()

        if self._source_selector.is_remote:
            src = self._source_selector.selected_source
            if not src:
                QMessageBox.warning(self, "Error", "No remote source selected.")
                return
            try:
                center_lat, center_lon = aoi.centroid_wgs84 if aoi else (lat, lon)
                source_path = resolve_remote_url(
                    src["key"], center_lat, center_lon,
                    year=self._source_selector.selected_year,
                )
            except Exception as exc:
                QMessageBox.warning(self, "Error", f"Failed to resolve remote URL:\n{exc}")
                return
        else:
            source_path = self._source_selector.selected_file
            if not source_path:
                QMessageBox.warning(
                    self, "Error",
                    "Please select a local raster file."
                )
                return

        selected_source = self._source_selector.selected_source
        self._last_params = {
            "lat": lat, "lon": lon, "radius_m": radius,
            "legend_key": legend_key, "source_path": source_path, "aoi": aoi,
            "source_key": selected_source.get("key") if selected_source else None,
            "source_year": self._source_selector.selected_year,
            "source_collection": selected_source.get("collection") if selected_source else None,
            "source_resolution": selected_source.get("resolution") if selected_source else None,
        }

        if aoi is not None:
            self._map_viewer.show_aoi(aoi, center=aoi.centroid_wgs84)
        else:
            gj = buffer_geojson(lon, lat, radius)
            self._map_viewer.show_buffer(lat, lon, radius, gj)

        # Run analysis in thread
        self._btn_analyze.setEnabled(False)
        self._status_bar.showMessage(t("status.analyzing"))

        self._thread = QThread()
        self._worker = AnalysisWorker(
            source_path, lon, lat, radius, legend_key, aoi=aoi
        )
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished.connect(self._on_analysis_done)
        self._worker.finished.connect(self._thread.quit)
        self._thread.finished.connect(lambda: setattr(self, "_thread", None))
        self._thread.start()

    def _on_analysis_done(self, result: AnalysisResult | Exception) -> None:
        self._btn_analyze.setEnabled(True)

        if isinstance(result, Exception):
            self._status_bar.showMessage(t("status.error", msg=str(result)))
            QMessageBox.critical(
                self, "Analysis Error",
                f"An error occurred:\n\n{result}\n\n{traceback.format_exc()}"
            )
            return

        self._last_result = result
        self._summary_panel.update_result(result)
        self._results_table.update_results(result)
        self._metrics_panel.update_metrics(result.landscape_metrics, result.pixel_area_m2)
        self._warnings_panel.update_warnings(
            result.quality_warnings, result.total_valid_pixels
        )

        # Update map with results
        lat = self._last_params["lat"]
        lon = self._last_params["lon"]
        radius = self._last_params["radius_m"]
        aoi = self._last_params.get("aoi")
        if aoi is not None:
            lat, lon = aoi.centroid_wgs84
            radius = math.sqrt(aoi.area_m2 / math.pi)
            self._map_viewer.show_results(
                lat, lon, radius, result.class_stats, aoi.to_geojson(),
                raster_data=result.raster_data,
                raster_valid_mask=result.raster_valid_mask,
                raster_transform=result.raster_transform,
                raster_crs=result.raster_crs,
            )
        else:
            gj = buffer_geojson(lon, lat, radius)
            self._map_viewer.show_results(
                lat, lon, radius, result.class_stats, gj,
                raster_data=result.raster_data,
                raster_valid_mask=result.raster_valid_mask,
                raster_transform=result.raster_transform,
                raster_crs=result.raster_crs,
            )

        self._status_bar.showMessage(
            f"{t('status.done')} — {t('status.pixels_analyzed', n=result.total_valid_pixels)}"
        )

    # ── Temporal analysis ─────────────────────────────────────────────────

    def _resolve_temporal_path(self, path: str, year: int, lat: float, lon: float) -> str:
        """Return a local path or resolve the selected remote catalog year."""
        if path:
            return path
        if not self._source_selector.is_remote:
            raise ValueError("Selecione os dois arquivos GeoTIFF ou ative uma fonte remota.")
        source = self._source_selector.selected_source
        if not source:
            raise ValueError("Nenhuma fonte remota selecionada.")
        validate_year(source["key"], year)
        return resolve_remote_url(source["key"], lat, lon, year=year)

    def _start_background_task(self, task, callback) -> bool:
        """Run a pure analysis callable without blocking the Qt event loop."""
        if self._active_task_thread is not None or (
            getattr(self, "_thread", None) is not None and self._thread.isRunning()
        ):
            QMessageBox.information(
                self, "Análise em andamento", "Aguarde a análise atual terminar."
            )
            return False
        thread = QThread(self)
        worker = TaskWorker(task)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(callback)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda: self._clear_background_task(thread))
        self._active_task_thread = thread
        self._active_task_worker = worker
        thread.start()
        return True

    def _clear_background_task(self, thread: QThread) -> None:
        if self._active_task_thread is thread:
            self._active_task_thread = None
            self._active_task_worker = None

    @staticmethod
    def _compute_temporal_transition(
        file1: str, file2: str, legend_key: str,
        lat: float, lon: float, radius: float, aoi: AOI | None,
    ) -> dict:
        legend_classes = load_legend_classes(legend_key)
        if aoi is not None:
            r1 = clip_raster_to_geometry(file1, aoi)
            r2 = clip_raster_to_geometry(file2, aoi)
        else:
            r1 = clip_raster_to_buffer(file1, lon, lat, radius)
            r2 = clip_raster_to_buffer(file2, lon, lat, radius)
        r1, r2 = align_raster_pair(r1, r2)
        m1 = r1.valid_mask_for_legend(legend_classes)
        m2 = r2.valid_mask_for_legend(legend_classes)
        cs1 = compute_class_statistics(r1.data, m1, r1.pixel_area_m2, legend_classes)
        cs2 = compute_class_statistics(r2.data, m2, r1.pixel_area_m2, legend_classes)
        transition = compute_transition_matrix(
            r1.data, r2.data, m1, m2, r1.pixel_area_m2, legend_classes
        )
        transition["metrics_t1"] = compute_landscape_metrics(
            cs1, r1.data, m1, r1.pixel_area_m2
        )
        transition["metrics_t2"] = compute_landscape_metrics(
            cs2, r2.data, m2, r1.pixel_area_m2
        )
        return transition

    @staticmethod
    def _compute_temporal_series(
        year_file_pairs: list, legend_key: str,
        lat: float, lon: float, radius: float, aoi: AOI | None,
    ) -> list[dict]:
        legend_classes = load_legend_classes(legend_key)
        series: list[dict] = []
        reference = None
        for year, file_path in sorted(year_file_pairs, key=lambda x: x[0]):
            if aoi is not None:
                raster = clip_raster_to_geometry(file_path, aoi)
            else:
                raster = clip_raster_to_buffer(file_path, lon, lat, radius)
            if reference is None:
                reference = raster
            else:
                raster = align_raster_to_reference(raster, reference)
            valid_mask = raster.valid_mask_for_legend(legend_classes)
            cs = compute_class_statistics(
                raster.data, valid_mask, reference.pixel_area_m2, legend_classes
            )
            lm = compute_landscape_metrics(
                cs, raster.data, valid_mask, reference.pixel_area_m2
            )
            series.append({"year": year, "class_stats": cs, "landscape_metrics": lm})
        return series

    def _run_temporal_analysis(
        self, file1: str, file2: str, year1: int, year2: int
    ) -> None:
        lat = self._lat_input.value()
        lon = self._lon_input.value()
        radius = self._radius_input.value()
        legend_key = self._source_selector.get_legend_key()
        aoi = self._aoi

        try:
            center_lat, center_lon = (
                aoi.centroid_wgs84 if aoi is not None else (lat, lon)
            )
            resolved1 = self._resolve_temporal_path(file1, year1, center_lat, center_lon)
            resolved2 = self._resolve_temporal_path(file2, year2, center_lat, center_lon)
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return

        self._status_bar.showMessage(t("status.analyzing"))
        self._start_background_task(
            lambda: self._compute_temporal_transition(
                resolved1, resolved2, legend_key, lat, lon, radius, aoi
            ),
            lambda result: self._on_temporal_done(result, year1, year2),
        )
        return

    def _on_temporal_done(self, result, year1: int, year2: int) -> None:
        if isinstance(result, Exception):
            self._status_bar.showMessage(t("status.error", msg=str(result)))
            QMessageBox.critical(self, "Error", str(result))
            return
        self._last_temporal = result
        self._last_temporal_years = (year1, year2)
        self._temporal_panel.update_results(result, year1, year2)
        self._status_bar.showMessage(t("status.done"))

    def _run_temporal_series(self, year_file_pairs: list) -> None:
        """Analyse N individual years and display a longitudinal coverage table."""
        legend_key = self._source_selector.get_legend_key()
        lat = self._lat_input.value()
        lon = self._lon_input.value()
        radius = self._radius_input.value()
        aoi = self._aoi

        try:
            center_lat, center_lon = (
                aoi.centroid_wgs84 if aoi is not None else (lat, lon)
            )
            resolved_pairs = [
                (
                    year,
                    self._resolve_temporal_path(path, int(year), center_lat, center_lon),
                )
                for year, path in year_file_pairs
            ]
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return

        self._status_bar.showMessage(t("status.analyzing"))
        self._start_background_task(
            lambda: self._compute_temporal_series(
                resolved_pairs, legend_key, lat, lon, radius, aoi
            ),
            self._on_temporal_series_done,
        )
        return

    # ── Multi-point comparison ────────────────────────────────────────────

    def _on_temporal_series_done(self, result) -> None:
        if isinstance(result, Exception):
            self._status_bar.showMessage(t("status.error", msg=str(result)))
            QMessageBox.critical(self, "Error", str(result))
            return
        self._last_temporal_series = result
        self._temporal_panel.update_series_results(result)
        self._status_bar.showMessage(t("status.done"))

    @staticmethod
    def _compute_aoi_comparison(
        areas: list[tuple[str, AOI, str]], legend_key: str,
    ) -> list[dict]:
        legend_classes = load_legend_classes(legend_key)
        results = []
        for label, aoi, source_path in areas:
            raster = clip_raster_to_geometry(source_path, aoi)
            valid_mask = raster.valid_mask_for_legend(legend_classes)
            class_stats = compute_class_statistics(
                raster.data, valid_mask, raster.pixel_area_m2, legend_classes
            )
            metrics = compute_landscape_metrics(
                class_stats, raster.data, valid_mask, raster.pixel_area_m2
            )
            results.append({
                "point_label": label,
                "class_stats": class_stats,
                "landscape_metrics": metrics,
                "geometry_area_m2": aoi.area_m2,
            })
        return results

    def _run_aoi_comparison(self, areas: list[tuple[str, AOI]]) -> None:
        legend_key = self._source_selector.get_legend_key()
        source = self._source_selector.selected_source
        if self._source_selector.is_remote:
            if not source:
                QMessageBox.warning(self, "Error", "Nenhuma fonte remota selecionada.")
                return
            year = self._source_selector.selected_year
            if year is None:
                QMessageBox.warning(self, "Error", "Selecione o ano da fonte remota.")
                return
            try:
                validate_year(source["key"], year)
            except ValueError as exc:
                QMessageBox.warning(self, "Error", str(exc))
                return
        elif self._source_selector.selected_file:
            year = None
        else:
            QMessageBox.warning(self, "Error", "Selecione um raster local ou fonte remota.")
            return

        resolved_areas = []
        try:
            for label, aoi in areas:
                if self._source_selector.is_remote:
                    lat, lon = aoi.centroid_wgs84
                    source_path = resolve_remote_url(
                        source["key"], lat, lon, year=year
                    )
                else:
                    source_path = self._source_selector.selected_file
                resolved_areas.append((label, aoi, source_path))
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return

        self._status_bar.showMessage(t("status.analyzing"))
        self._start_background_task(
            lambda: self._compute_aoi_comparison(resolved_areas, legend_key),
            self._on_aoi_comparison_done,
        )

    def _on_aoi_comparison_done(self, result) -> None:
        if isinstance(result, Exception):
            self._status_bar.showMessage(t("status.error", msg=str(result)))
            QMessageBox.critical(self, "Error", str(result))
            return
        self._aoi_compare_results.update_aoi_comparison(result)
        # Reuse the comparison export/map pipeline while retaining the exact
        # polygon area in each result record.
        self._last_compare = result
        self._last_compare_points = None
        areas = self._aoi_compare_panel.areas
        exact_areas = [
            {"label": result[index]["point_label"], "aoi": area}
            for index, (_, area) in enumerate(areas)
            if index < len(result)
        ]
        if exact_areas:
            self._compare_map_viewer.show_compare_aois(exact_areas)
        self._status_bar.showMessage(t("status.done"))

    @staticmethod
    def _compute_comparison(source_path: str, legend_key: str, points: list[tuple]) -> tuple[list[dict], list[dict]]:
        legend_classes = load_legend_classes(legend_key)
        results: list[dict] = []
        map_points: list[dict] = []
        for i, (label, lat, lon, radius) in enumerate(points):
            raster = clip_raster_to_buffer(source_path, lon, lat, radius)
            valid_mask = raster.valid_mask_for_legend(legend_classes)
            cs = compute_class_statistics(
                raster.data, valid_mask, raster.pixel_area_m2, legend_classes
            )
            lm = compute_landscape_metrics(
                cs, raster.data, valid_mask, raster.pixel_area_m2
            )
            results.append({
                "point_label": label,
                "class_stats": cs,
                "landscape_metrics": lm,
            })
            map_points.append({
                "label": label, "lat": lat, "lon": lon, "radius_m": radius,
            })
        return results, map_points

    @staticmethod
    def _compute_comparison_sources(
        entries: list[tuple[str, float, float, float, str]], legend_key: str,
    ) -> tuple[list[dict], list[dict]]:
        legend_classes = load_legend_classes(legend_key)
        results, map_points = [], []
        for label, lat, lon, radius, source_path in entries:
            raster = clip_raster_to_buffer(source_path, lon, lat, radius)
            valid_mask = raster.valid_mask_for_legend(legend_classes)
            class_stats = compute_class_statistics(
                raster.data, valid_mask, raster.pixel_area_m2, legend_classes
            )
            metrics = compute_landscape_metrics(
                class_stats, raster.data, valid_mask, raster.pixel_area_m2
            )
            results.append({
                "point_label": label,
                "class_stats": class_stats,
                "landscape_metrics": metrics,
            })
            map_points.append({"label": label, "lat": lat, "lon": lon, "radius_m": radius})
        return results, map_points

    def _on_comparison_done(self, result) -> None:
        if isinstance(result, Exception):
            self._status_bar.showMessage(t("status.error", msg=str(result)))
            QMessageBox.critical(self, "Error", str(result))
            return
        results, map_points = result
        self._last_compare = results
        self._last_compare_points = map_points
        self._compare_panel.update_results(results)
        self._compare_map_viewer.show_compare_points(map_points)
        self._status_bar.showMessage(t("status.done"))

    def _run_comparison(self, points: list) -> None:
        """points: list[ComparePoint] with .name .lat .lon .radius"""
        source_path = self._source_selector.selected_file
        legend_key = self._source_selector.get_legend_key()
        source = self._source_selector.selected_source

        if not source_path and not self._source_selector.is_remote:
            QMessageBox.warning(self, "Error", "Selecione um raster local ou fonte remota.")
            return

        normalized_points = []
        for i, point in enumerate(points):
            if isinstance(point, tuple):
                lat, lon, radius = point
                label = f"P{i + 1}"
            else:
                lat, lon, radius = point.lat, point.lon, point.radius
                label = point.name
            normalized_points.append((label, float(lat), float(lon), float(radius)))
        if self._source_selector.is_remote:
            year = self._source_selector.selected_year
            if source is None or year is None:
                QMessageBox.warning(self, "Error", "Selecione a fonte e o ano.")
                return
            try:
                validate_year(source["key"], year)
                entries = [
                    (*point, resolve_remote_url(source["key"], point[1], point[2], year=year))
                    for point in normalized_points
                ]
            except Exception as exc:
                QMessageBox.critical(self, "Error", str(exc))
                return
            task = lambda: self._compute_comparison_sources(entries, legend_key)
        else:
            task = lambda: self._compute_comparison(source_path, legend_key, normalized_points)
        self._status_bar.showMessage(t("status.analyzing"))
        self._start_background_task(task, self._on_comparison_done)
    # ── Exports ───────────────────────────────────────────────────────────

    def _export_csv(self) -> None:
        if not self._last_result:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", "luma_results.csv", "CSV (*.csv)"
        )
        if not path:
            return
        result = self._last_result
        with open(path, "w", encoding="utf-8") as f:
            f.write("class_id,class_name,pixels,area_m2,area_km2,area_ha,percentage,num_patches,largest_patch_m2,color\n")
            for cs in result.class_stats:
                f.write(
                    f"{cs.class_id},{cs.class_name},{cs.pixel_count},"
                    f"{cs.area_m2:.2f},{cs.area_m2/1e6:.6f},{cs.area_m2/10000:.4f},"
                    f"{cs.percentage:.2f},{cs.num_patches},"
                    f"{cs.largest_patch_area_m2:.2f},{cs.color}\n"
                )
        self._status_bar.showMessage(f"CSV exported to {path}")

    def _export_xlsx(self) -> None:
        """Export all available analyses to a multi-sheet .xlsx file."""
        if not (
            self._last_result or self._last_temporal
            or self._last_temporal_series or self._last_compare
        ):
            QMessageBox.warning(self, "Error", "No analysis results to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "luma_results.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(
                self, "Error",
                "openpyxl is required for Excel export. Install with: pip install openpyxl",
            )
            return

        wb = openpyxl.Workbook()
        # Remove default sheet (we'll create our own)
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        def _write_header(ws, row_idx: int, headers: list[str]) -> None:
            for col, val in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # ── Sheet: Classes (single analysis) ────────────────────────────
        if self._last_result:
            ws = wb.create_sheet("Classes")
            _write_header(ws, 1, [
                "class_id", "class_name", "pixels", "area_m2",
                "area_km2", "area_ha", "percentage", "num_patches",
                "largest_patch_m2", "color",
            ])
            for i, cs in enumerate(self._last_result.class_stats, 2):
                ws.cell(row=i, column=1, value=cs.class_id)
                ws.cell(row=i, column=2, value=cs.class_name)
                ws.cell(row=i, column=3, value=cs.pixel_count)
                ws.cell(row=i, column=4, value=round(cs.area_m2, 2))
                ws.cell(row=i, column=5, value=round(cs.area_m2 / 1e6, 6))
                ws.cell(row=i, column=6, value=round(cs.area_m2 / 10_000, 4))
                ws.cell(row=i, column=7, value=round(cs.percentage, 2))
                ws.cell(row=i, column=8, value=cs.num_patches)
                ws.cell(row=i, column=9, value=round(cs.largest_patch_area_m2, 2))
                ws.cell(row=i, column=10, value=cs.color)

            # Metrics sheet
            ws_m = wb.create_sheet("Metrics")
            _write_header(ws_m, 1, ["metric", "value"])
            m = self._last_result.landscape_metrics
            rows = [
                ("shannon_diversity", m.shannon_diversity),
                ("isa_index", m.isa_index),
                ("simpson_diversity", m.simpson_diversity),
                ("dominance", m.dominance),
                ("evenness", m.evenness),
                ("total_patches", m.total_patches),
                ("patch_density", m.patch_density),
                ("largest_patch_index", m.largest_patch_index),
                ("edge_density", m.edge_density),
                ("effective_mesh_size", m.effective_mesh_size),
                ("aggregation_index", m.aggregation_index),
                ("contagion", m.contagion),
                ("mean_shape_index", m.mean_shape_index),
            ]
            for i, (k, v) in enumerate(rows, 2):
                ws_m.cell(row=i, column=1, value=k)
                ws_m.cell(row=i, column=2, value=v)

        # ── Sheet: Temporal ─────────────────────────────────────────────
        if self._last_temporal:
            ws = wb.create_sheet("Temporal")
            td = self._last_temporal
            classes = td["classes"]
            legend = td["legend"]
            matrix = td["matrix"]
            net_change = td["net_change"]
            y1, y2 = self._last_temporal_years
            class_names = [legend.get(c, {}).get("name", f"Class {c}") for c in classes]

            ws.cell(row=1, column=1, value=f"Transition matrix (km²) — {y1} → {y2}")
            ws.cell(row=1, column=1).font = Font(bold=True)
            _write_header(ws, 2, [f"from \\ to ({y1} → {y2})"] + class_names)
            for i, c1 in enumerate(classes):
                ws.cell(row=3 + i, column=1, value=class_names[i])
                ws.cell(row=3 + i, column=1).font = Font(bold=True)
                for j, c2 in enumerate(classes):
                    ws.cell(row=3 + i, column=2 + j, value=round(matrix[c1].get(c2, 0) / 1e6, 4))

            base = 3 + len(classes) + 2
            ws.cell(row=base, column=1, value="Persistence (%)").font = Font(bold=True)
            ws.cell(row=base, column=2, value=td["persistence"])
            ws.cell(row=base + 2, column=1, value="Net change by class (km²)").font = Font(bold=True)
            _write_header(ws, base + 3, ["class", "net_change_km2"])
            for i, c in enumerate(classes):
                ws.cell(row=base + 4 + i, column=1, value=class_names[i])
                ws.cell(row=base + 4 + i, column=2, value=round(net_change.get(c, 0) / 1e6, 4))
            if "metrics_t1" in td and "metrics_t2" in td:
                isa_base = base + 4 + len(classes) + 2
                ws.cell(row=isa_base, column=1, value=f"ISA {y1} (%)").font = Font(bold=True)
                ws.cell(row=isa_base, column=2, value=td["metrics_t1"].isa_index)
                ws.cell(row=isa_base + 1, column=1, value=f"ISA {y2} (%)").font = Font(bold=True)
                ws.cell(row=isa_base + 1, column=2, value=td["metrics_t2"].isa_index)
                ws.cell(row=isa_base + 2, column=1, value="ISA delta (p.p.)").font = Font(bold=True)
                ws.cell(
                    row=isa_base + 2,
                    column=2,
                    value=round(td["metrics_t2"].isa_index - td["metrics_t1"].isa_index, 2),
                )

        # ── Sheet: Compare points (points = rows, classes + metrics = cols)
        if self._last_temporal_series:
            ws = wb.create_sheet("Temporal Series")
            series = self._last_temporal_series
            years = [entry["year"] for entry in series]
            _write_header(ws, 1, ["metric"] + [str(y) for y in years])
            ws.cell(row=2, column=1, value="isa_index")
            for col, entry in enumerate(series, 2):
                ws.cell(row=2, column=col, value=entry["landscape_metrics"].isa_index)

            all_classes: list[str] = []
            seen = set()
            for entry in series:
                for cs in entry["class_stats"]:
                    if cs.class_name not in seen:
                        all_classes.append(cs.class_name)
                        seen.add(cs.class_name)
            for row, cls_name in enumerate(all_classes, 3):
                ws.cell(row=row, column=1, value=f"%_{cls_name}")
                for col, entry in enumerate(series, 2):
                    pct = next(
                        (cs.percentage for cs in entry["class_stats"] if cs.class_name == cls_name),
                        0.0,
                    )
                    ws.cell(row=row, column=col, value=round(pct, 2))

        if self._last_compare:
            ws = wb.create_sheet("Compare")
            all_classes: list[str] = []
            seen = set()
            for r in self._last_compare:
                for cs in r["class_stats"]:
                    if cs.class_name not in seen:
                        all_classes.append(cs.class_name)
                        seen.add(cs.class_name)

            metric_cols = [
                ("shannon_diversity", lambda m: m.shannon_diversity),
                ("isa_index", lambda m: m.isa_index),
                ("simpson_diversity", lambda m: m.simpson_diversity),
                ("evenness", lambda m: m.evenness),
                ("total_patches", lambda m: m.total_patches),
                ("patch_density", lambda m: m.patch_density),
                ("largest_patch_index", lambda m: m.largest_patch_index),
                ("largest_patch_area_m2", lambda m: m.largest_patch_area_m2),
                ("smallest_patch_area_m2", lambda m: m.smallest_patch_area_m2),
                ("mean_patch_area_m2", lambda m: m.mean_patch_area_m2),
                ("aggregation_index", lambda m: m.aggregation_index),
                ("contagion", lambda m: m.contagion),
                ("mean_shape_index", lambda m: m.mean_shape_index),
            ]

            headers = ["point"] + [f"%_{c}" for c in all_classes] + [n for n, _ in metric_cols]
            _write_header(ws, 1, headers)

            for i, r in enumerate(self._last_compare, 2):
                ws.cell(row=i, column=1, value=r["point_label"])
                for c_off, cls_name in enumerate(all_classes):
                    pct = 0.0
                    for cs in r["class_stats"]:
                        if cs.class_name == cls_name:
                            pct = cs.percentage
                            break
                    ws.cell(row=i, column=2 + c_off, value=round(pct, 2))
                lm = r["landscape_metrics"]
                for m_off, (_, acc) in enumerate(metric_cols):
                    col = 2 + len(all_classes) + m_off
                    ws.cell(row=i, column=col, value=acc(lm))

        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.critical(self, "Error",
                "File is open in another program. Close it and try again.")
            return

        self._status_bar.showMessage(f"Excel exported to {path}")

    def _export_comparison_json(self) -> None:
        import json
        path, _ = QFileDialog.getSaveFileName(
            self, "Export JSON", "luma_comparison.json", "JSON (*.json)"
        )
        if not path:
            return
        payload = []
        for item in self._last_compare or []:
            payload.append({
                "label": item.get("point_label", ""),
                "geometry_area_m2": item.get("geometry_area_m2"),
                "classes": [
                    {
                        "id": stat.class_id,
                        "name": stat.class_name,
                        "pixels": stat.pixel_count,
                        "area_m2": stat.area_m2,
                        "percentage": stat.percentage,
                    }
                    for stat in item.get("class_stats", [])
                ],
                "landscape_metrics": {
                    "total_patches": item["landscape_metrics"].total_patches,
                    "shannon_diversity": item["landscape_metrics"].shannon_diversity,
                    "largest_patch_index": item["landscape_metrics"].largest_patch_index,
                },
            })
        with open(path, "w", encoding="utf-8") as handle:
            json.dump({"comparison": payload}, handle, ensure_ascii=False, indent=2)
        self._status_bar.showMessage(f"JSON exported to {path}")

    def _export_json(self) -> None:
        if not self._last_result:
            if self._last_compare:
                self._export_comparison_json()
            return
        import json
        path, _ = QFileDialog.getSaveFileName(
            self, "Export JSON", "luma_results.json", "JSON (*.json)"
        )
        if not path:
            return
        result = self._last_result
        data = {
            "parameters": serializable_parameters(self._last_params),
            "total_area_m2": result.total_area_m2,
            "total_pixels": result.total_valid_pixels,
            "source": result.source_name,
            "accuracy": result.source_accuracy,
            "classes": [
                {
                    "id": cs.class_id,
                    "name": cs.class_name,
                    "pixels": cs.pixel_count,
                    "area_m2": cs.area_m2,
                    "percentage": cs.percentage,
                    "num_patches": cs.num_patches,
                    "largest_patch_m2": cs.largest_patch_area_m2,
                }
                for cs in result.class_stats
            ],
            "landscape_metrics": {
                "shannon_diversity": result.landscape_metrics.shannon_diversity,
                "isa_index": result.landscape_metrics.isa_index,
                "simpson_diversity": result.landscape_metrics.simpson_diversity,
                "dominance": result.landscape_metrics.dominance,
                "evenness": result.landscape_metrics.evenness,
                "total_patches": result.landscape_metrics.total_patches,
                "patch_density": result.landscape_metrics.patch_density,
                "largest_patch_index": result.landscape_metrics.largest_patch_index,
                "edge_density": result.landscape_metrics.edge_density,
                "effective_mesh_size": result.landscape_metrics.effective_mesh_size,
                "aggregation_index": result.landscape_metrics.aggregation_index,
                "contagion": result.landscape_metrics.contagion,
                "mean_shape_index": result.landscape_metrics.mean_shape_index,
            },
            "warnings": result.quality_warnings,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        self._status_bar.showMessage(f"JSON exported to {path}")

    def _export_pdf(self) -> None:
        has_any = (
            self._last_result or self._last_temporal
            or self._last_temporal_series or self._last_compare
        )
        if not has_any:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export PDF", "luma_report.pdf", "PDF (*.pdf)"
        )
        if not path:
            return
        try:
            from luma.output.report import generate_pdf_report
            # Capture compare map if available
            compare_map_img: bytes | None = None
            if self._last_compare:
                pix = self._compare_map_viewer.grab_map()
                if pix and not pix.isNull():
                    import io
                    from PySide6.QtCore import QBuffer, QIODevice
                    buf = QBuffer()
                    buf.open(QIODevice.OpenModeFlag.WriteOnly)
                    pix.save(buf, "PNG")
                    compare_map_img = bytes(buf.data())

            generate_pdf_report(
                path=path,
                result=self._last_result,
                params=self._last_params,
                lang=get_language(),
                temporal_data=self._last_temporal,
                temporal_years=self._last_temporal_years,
                compare_data=self._last_compare,
                temporal_series=self._last_temporal_series,
                compare_map_img=compare_map_img,
            )
            self._status_bar.showMessage(f"PDF exported to {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"PDF export failed:\n{exc}")

    def _comparison_points(self):
        return list(self._last_compare_points or [])

    def _on_compare_open_map(self, points: list) -> None:
        map_points = [
            {"label": point.name, "lat": point.lat, "lon": point.lon, "radius_m": point.radius}
            for point in points
        ]
        self._compare_map_viewer.show_compare_points(map_points)
        self._tabs.setCurrentIndex(3)

    def _on_compare_bulk_download(self, target_dir: str) -> None:
        """Export WGS-84 vectors and the current comparison map."""
        if not self._last_compare_points:
            QMessageBox.warning(self, "", t("compare_extra.bulk_download_no_data"))
            return
        points = self._comparison_points()
        if not points:
            return
        try:
            output_dir = Path(target_dir)
            outputs = export_points_buffers(output_dir, points)
            map_path = output_dir / "compare_map.tif"
            class_stats = self._last_compare[0].get("class_stats", [])
            if not self._compare_map_viewer.export_tiff_with_legend(
                str(map_path), class_stats=class_stats, title=t("tabs.compare_map")
            ):
                QMessageBox.warning(self, "Error", t("menu.compare_tiff_no_map"))
            else:
                outputs["map_tif"] = map_path
            self._status_bar.showMessage(
                t("compare_extra.bulk_download_done", path=str(output_dir))
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"{t('menu.export_geospatial')}:\n{exc}")

    def _on_compare_map_tiff(self, path: str) -> None:
        class_stats = self._last_compare[0].get("class_stats", []) if self._last_compare else []
        if self._compare_map_viewer.export_tiff_with_legend(
            path, class_stats=class_stats, title=t("tabs.compare_map")
        ):
            self._status_bar.showMessage(f"TIFF exported to {path}")
        else:
            QMessageBox.warning(self, "Error", t("menu.compare_tiff_no_map"))

    def _export_geospatial_bundle(self) -> None:
        if not self._last_compare_points:
            QMessageBox.information(self, "", t("compare_extra.bulk_download_no_data"))
            return
        target_dir = QFileDialog.getExistingDirectory(
            self, t("compare_extra.bulk_download_dir")
        )
        if target_dir:
            self._on_compare_bulk_download(target_dir)

    def _export_chart(self) -> None:
        options: list[tuple[str, str]] = []
        if self._last_temporal_series:
            options.append((t("menu.chart_temporal"), "temporal"))
        if self._last_compare:
            options.append((t("menu.chart_compare"), "compare"))
        if not options:
            QMessageBox.information(self, "", t("menu.chart_no_data"))
            return
        labels = [label for label, _ in options]
        selected, ok = QInputDialog.getItem(
            self, t("menu.export_chart"), t("menu.chart_choose"), labels, 0, False
        )
        if not ok:
            return
        key = dict(options)[selected]
        path, _ = QFileDialog.getSaveFileName(
            self, t("menu.export_chart"), "luma_chart.png",
            "PNG (*.png);;SVG (*.svg);;PDF (*.pdf)",
        )
        if not path:
            return
        try:
            if key == "temporal":
                figure = build_temporal_series_figure(
                    self._last_temporal_series,
                    getattr(self._temporal_panel, "chart_type", "bar"),
                )
            else:
                figure = build_compare_gradient_figure(
                    self._last_compare,
                    getattr(self._compare_panel, "gradient_metric", "isa"),
                )
            save_figure(figure, path)
            self._status_bar.showMessage(f"Chart exported to {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"{t('menu.export_chart')}:\n{exc}")

    def _export_compare_tiff(self) -> None:
        """Save the comparison map as a TIFF with its class legend."""
        if not self._last_compare:
            QMessageBox.information(self, "", t("menu.compare_tiff_no_data"))
            return
        path, _ = QFileDialog.getSaveFileName(
            self, t("menu.export_compare_tiff"), "luma_compare_map.tif",
            "TIFF (*.tif *.tiff)",
        )
        if path:
            self._on_compare_map_tiff(path)

    # ── Dialogs ───────────────────────────────────────────────────────────

    def _show_about(self) -> None:
        dlg = AboutDialog(self)
        dlg.exec()

    def _open_settings(self) -> None:
        dlg = SettingsDialog(self)
        dlg.language_changed.connect(self._on_language_changed)
        dlg.exec()

    def _on_language_changed(self, lang: str) -> None:
        set_language(lang)
        self.setWindowTitle(t("app.title"))
        self._btn_analyze.setText(t("input.analyze"))

        # Update input labels and help bubbles
        self._lbl_lat.setText(t("input.latitude"))
        self._lbl_lon.setText(t("input.longitude"))
        self._lbl_rad.setText(t("input.radius"))
        self._tip_lat.set_tip(t("tips.latitude"))
        self._tip_lon.set_tip(t("tips.longitude"))
        self._tip_rad.set_tip(t("tips.radius"))

        self._source_selector.refresh_texts()
        self._objective_selector.refresh_texts()
        self._aoi_widget.refresh_texts()
        self._summary_panel.refresh_texts()
        self._results_table.refresh_texts()
        self._metrics_panel.refresh_texts()
        self._temporal_panel.refresh_texts()
        self._compare_panel.refresh_texts()
        self._aoi_compare_panel.refresh_texts()
        self._aoi_compare_results.refresh_texts()
        self._btn_open_inputs.setText(t("input.parameters"))
        self._btn_compact_results.setText(t("input.view_results"))
        self._tabs.setTabText(0, t("tabs.single"))
        self._tabs.setTabText(1, t("tabs.temporal"))
        self._tabs.setTabText(2, t("tabs.compare"))
        self._tabs.setTabText(3, t("tabs.compare_map"))
        self._tabs.setTabText(4, t("tabs.compare_aois"))
        self._result_tabs.setTabText(0, t("tabs.coverage"))
        self._result_tabs.setTabText(1, t("tabs.metrics"))
        self._status_bar.showMessage(t("status.ready"))
        self._setup_menu()

    def _save_project(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, t("menu.save_project"), "luma_project.luma.json",
            "LUMA Project (*.luma.json);;JSON (*.json)",
        )
        if not path:
            return
        source = self._source_selector.selected_source
        params = {
            "lat": self._lat_input.value(),
            "lon": self._lon_input.value(),
            "radius_m": self._radius_input.value(),
            "aoi": self._aoi,
            "source_file": self._source_selector.selected_file,
        }
        payload = build_project(
            params,
            source_key=source.get("key") if source else None,
            source_year=self._source_selector.selected_year,
            legend_key=self._source_selector.get_legend_key(),
        )
        try:
            save_project(path, payload)
        except OSError as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return
        self._status_bar.showMessage(f"Projeto salvo em {path}")

    def _open_project(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, t("menu.open_project"), "",
            "LUMA Project (*.luma.json);;JSON (*.json)",
        )
        if not path:
            return
        try:
            payload = load_project(path)
            params = payload["parameters"]
            source = payload["source"]
            self._lat_input.setValue(float(params.get("lat", self._lat_input.value())))
            self._lon_input.setValue(float(params.get("lon", self._lon_input.value())))
            self._radius_input.setValue(float(params.get("radius_m", self._radius_input.value())))
            aoi_payload = params.get("aoi")
            if aoi_payload:
                aoi = AOI.from_geojson(
                    aoi_payload["geometry"], crs=aoi_payload.get("crs")
                )
                self._aoi_widget.set_aoi(aoi)
            else:
                self._aoi_widget.set_aoi(None)
            self._source_selector.apply_project(
                source_key=source.get("key"),
                source_year=source.get("year"),
                legend_key=source.get("legend"),
                source_file=params.get("source_file"),
            )
            self._last_params = dict(params)
        except (ValueError, KeyError, TypeError) as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return
        self._status_bar.showMessage(f"Projeto aberto: {path}")
