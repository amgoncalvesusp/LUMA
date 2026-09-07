"""Multi-point comparison panel with three input modes:

1. Manual — form-style rows with Name/Lat/Lon/Radius.
2. Paste  — paste a CSV/TSV/semicolon-separated table, then map columns.
3. Excel  — load an .xlsx/.csv file, pick a sheet, then map columns.
"""

from __future__ import annotations

import csv
import io
import math
import os
from dataclasses import dataclass

from PySide6.QtCore import Signal, Qt, QSize
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QPushButton,
    QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QComboBox, QLineEdit, QPlainTextEdit, QStackedWidget, QRadioButton,
    QButtonGroup, QCheckBox, QMessageBox, QFrame, QSizePolicy,
)

from luma.i18n.translator import t
from luma.gui.widgets.coord_input import LatitudeInput, LongitudeInput, RadiusInput

try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as _FigCanvas
    _MPL_OK = True
except Exception:
    _MPL_OK = False


# ── Data model ───────────────────────────────────────────────────────────────

@dataclass
class ComparePoint:
    name: str
    lat: float
    lon: float
    radius: float


# ── Manual input row ─────────────────────────────────────────────────────────

class PointRow(QWidget):
    """A single row for manually entering a comparison point."""

    removed = Signal(object)  # emits self

    def __init__(self, index: int, parent: QWidget | None = None):
        super().__init__(parent)
        self.index = index
        layout = QGridLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText(t("compare.auto_name", n=index + 1))
        self.name_input.setMinimumWidth(120)

        self.lat_input = LatitudeInput()
        self.lon_input = LongitudeInput()
        self.radius_input = RadiusInput()

        self.lbl_name = QLabel(t("compare.col_name") + ":")
        self.lbl_lat = QLabel("Lat:")
        self.lbl_lon = QLabel("Lon:")
        self.lbl_r = QLabel("R:")

        layout.addWidget(self.lbl_name, 0, 0)
        layout.addWidget(self.name_input, 0, 1, 1, 4)
        layout.addWidget(self.lbl_lat, 1, 0)
        layout.addWidget(self.lat_input, 1, 1)
        layout.addWidget(self.lbl_lon, 1, 2)
        layout.addWidget(self.lon_input, 1, 3)
        layout.addWidget(self.lbl_r, 1, 4)
        layout.addWidget(self.radius_input, 1, 5)
        layout.setColumnStretch(1, 1)
        layout.setColumnStretch(3, 1)
        layout.setColumnStretch(5, 1)

        self.btn_remove = QPushButton("✕")
        self.btn_remove.setFixedSize(32, 32)
        self.btn_remove.setAccessibleName(t("input.remove_point"))
        self.btn_remove.setToolTip(t("input.remove_point"))
        self.btn_remove.setStyleSheet("color: #a93226; font-weight: bold;")
        self.btn_remove.clicked.connect(lambda: self.removed.emit(self))
        layout.addWidget(self.btn_remove, 0, 5, alignment=Qt.AlignmentFlag.AlignRight)

    def get_point(self) -> ComparePoint:
        name = self.name_input.text().strip() or t("compare.auto_name", n=self.index + 1)
        return ComparePoint(
            name=name,
            lat=self.lat_input.value(),
            lon=self.lon_input.value(),
            radius=self.radius_input.value(),
        )

    def refresh_texts(self) -> None:
        self.lbl_name.setText(t("compare.col_name") + ":")
        self.name_input.setPlaceholderText(t("compare.auto_name", n=self.index + 1))
        self.btn_remove.setAccessibleName(t("input.remove_point"))
        self.btn_remove.setToolTip(t("input.remove_point"))


# ── Table-based input (paste / excel share this) ─────────────────────────────

class _TableInput(QWidget):
    """A preview table plus column-mapping dropdowns.

    Subclasses populate the raw preview via `set_rows(headers, rows)`.
    """

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._headers: list[str] = []
        self._rows: list[list[str]] = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Preview
        self.lbl_preview = QLabel(t("compare.preview"))
        self.lbl_preview.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.lbl_preview)

        self.preview = QTableWidget()
        self.preview.setMaximumHeight(160)
        self.preview.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.preview)

        # Column mapping
        self.lbl_mapping = QLabel(t("compare.column_mapping"))
        self.lbl_mapping.setStyleSheet("font-weight: bold; margin-top: 4px;")
        layout.addWidget(self.lbl_mapping)

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        self.cmb_name = QComboBox()
        self.cmb_lat = QComboBox()
        self.cmb_lon = QComboBox()
        self.cmb_radius = QComboBox()

        self.lbl_col_name = QLabel(t("compare.col_name") + ":")
        self.lbl_col_lat = QLabel(t("compare.col_lat") + ":")
        self.lbl_col_lon = QLabel(t("compare.col_lon") + ":")
        self.lbl_col_radius = QLabel(t("compare.col_radius") + ":")

        grid.addWidget(self.lbl_col_name, 0, 0)
        grid.addWidget(self.cmb_name, 0, 1)
        grid.addWidget(self.lbl_col_lat, 0, 2)
        grid.addWidget(self.cmb_lat, 0, 3)
        grid.addWidget(self.lbl_col_lon, 1, 0)
        grid.addWidget(self.cmb_lon, 1, 1)
        grid.addWidget(self.lbl_col_radius, 1, 2)
        grid.addWidget(self.cmb_radius, 1, 3)
        layout.addLayout(grid)

        self.hint = QLabel(t("compare.name_col_hint"))
        self.hint.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        self.hint.setWordWrap(True)
        layout.addWidget(self.hint)

    def set_rows(self, headers: list[str], rows: list[list[str]]) -> None:
        self._headers = headers
        self._rows = rows

        # Fill preview
        self.preview.setColumnCount(len(headers))
        self.preview.setHorizontalHeaderLabels(headers)
        self.preview.setRowCount(min(len(rows), 8))
        for r, row in enumerate(rows[:8]):
            for c, val in enumerate(row):
                self.preview.setItem(r, c, QTableWidgetItem(str(val)))
        self.preview.resizeColumnsToContents()

        # Populate combo boxes
        for cmb in (self.cmb_name, self.cmb_lat, self.cmb_lon, self.cmb_radius):
            cmb.clear()
            cmb.addItem(t("compare.col_none"), -1)
            for i, h in enumerate(headers):
                cmb.addItem(h, i)

        # Best-effort auto-detect
        def _pick(cmb: QComboBox, candidates: list[str]) -> None:
            for i, h in enumerate(headers):
                hl = h.lower().strip()
                for c in candidates:
                    if c in hl:
                        cmb.setCurrentIndex(i + 1)  # +1 for None
                        return

        _pick(self.cmb_name, ["nome", "name", "label", "ponto", "id"])
        _pick(self.cmb_lat, ["lat"])
        _pick(self.cmb_lon, ["lon", "lng"])
        _pick(self.cmb_radius, ["raio", "radius", "buffer"])

    def get_mapped_points(self, fallback_radius: float) -> list[ComparePoint]:
        """Build ComparePoint list from current mapping. Raises ValueError on bad rows."""
        idx_name = self.cmb_name.currentData()
        idx_lat = self.cmb_lat.currentData()
        idx_lon = self.cmb_lon.currentData()
        idx_radius = self.cmb_radius.currentData()

        if idx_lat is None or idx_lat < 0 or idx_lon is None or idx_lon < 0:
            raise ValueError("Latitude/Longitude columns must be mapped.")

        out: list[ComparePoint] = []
        for i, row in enumerate(self._rows):
            try:
                lat = _to_float(row[idx_lat])
                lon = _to_float(row[idx_lon])
            except (ValueError, IndexError):
                raise ValueError(f"Linha {i + 1}: latitude/longitude inválidas.")
            if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                raise ValueError(f"Linha {i + 1}: latitude/longitude fora dos limites.")

            if idx_radius is not None and idx_radius >= 0:
                try:
                    r_val = _to_float(row[idx_radius])
                except (ValueError, IndexError):
                    raise ValueError(f"Linha {i + 1}: raio inválido.")
            else:
                r_val = fallback_radius
            if not math.isfinite(r_val) or r_val <= 0:
                raise ValueError(f"Linha {i + 1}: o raio deve ser maior que zero.")

            if idx_name is not None and idx_name >= 0:
                try:
                    name = str(row[idx_name]).strip()
                except IndexError:
                    name = ""
            else:
                name = ""
            if not name:
                name = t("compare.auto_name", n=len(out) + 1)

            out.append(ComparePoint(name=name, lat=lat, lon=lon, radius=r_val))
        return out

    def refresh_texts(self) -> None:
        self.lbl_preview.setText(t("compare.preview"))
        self.lbl_mapping.setText(t("compare.column_mapping"))
        self.lbl_col_name.setText(t("compare.col_name") + ":")
        self.lbl_col_lat.setText(t("compare.col_lat") + ":")
        self.lbl_col_lon.setText(t("compare.col_lon") + ":")
        self.lbl_col_radius.setText(t("compare.col_radius") + ":")
        self.hint.setText(t("compare.name_col_hint"))
        # Update "None" entries without losing selection
        for cmb in (self.cmb_name, self.cmb_lat, self.cmb_lon, self.cmb_radius):
            if cmb.count() > 0:
                cmb.setItemText(0, t("compare.col_none"))


# ── Paste mode ───────────────────────────────────────────────────────────────

class _PasteInput(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.hint = QLabel(t("compare.paste_hint"))
        self.hint.setWordWrap(True)
        self.hint.setStyleSheet("color: #555;")
        layout.addWidget(self.hint)

        self.text = QPlainTextEdit()
        self.text.setPlaceholderText(
            "name,lat,lon,radius\nSite A,-21.78,-48.17,5000\nSite B,-22.01,-47.90,5000"
        )
        self.text.setMaximumHeight(140)
        self.text.textChanged.connect(self._on_text_changed)
        layout.addWidget(self.text)

        self.table = _TableInput()
        layout.addWidget(self.table)

    def _on_text_changed(self) -> None:
        raw = self.text.toPlainText().strip()
        if not raw:
            self.table.set_rows([], [])
            return
        headers, rows = _parse_delimited(raw)
        self.table.set_rows(headers, rows)

    def get_points(self, fallback_radius: float) -> list[ComparePoint]:
        return self.table.get_mapped_points(fallback_radius)

    def refresh_texts(self) -> None:
        self.hint.setText(t("compare.paste_hint"))
        self.table.refresh_texts()


# ── Excel/CSV-file mode ──────────────────────────────────────────────────────

class _ExcelInput(QWidget):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._file_path: str | None = None
        self._sheets: dict[str, tuple[list[str], list[list[str]]]] = {}

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        top = QHBoxLayout()
        self.btn_browse = QPushButton(t("compare.excel_select"))
        self.btn_browse.clicked.connect(self._on_browse)
        top.addWidget(self.btn_browse)

        self.lbl_path = QLabel("")
        self.lbl_path.setStyleSheet("color: #555; font-style: italic;")
        top.addWidget(self.lbl_path, 1)
        layout.addLayout(top)

        sheet_row = QHBoxLayout()
        self.lbl_sheet = QLabel(t("compare.excel_sheet") + ":")
        self.cmb_sheet = QComboBox()
        self.cmb_sheet.currentIndexChanged.connect(self._on_sheet_changed)
        sheet_row.addWidget(self.lbl_sheet)
        sheet_row.addWidget(self.cmb_sheet, 1)
        layout.addLayout(sheet_row)

        self.table = _TableInput()
        layout.addWidget(self.table)

    def _on_browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, t("compare.excel_select"), "",
            "Excel/CSV (*.xlsx *.xls *.csv);;All files (*)",
        )
        if not path:
            return
        try:
            self._sheets = _read_spreadsheet(path)
        except Exception as exc:
            QMessageBox.warning(self, "Error", t("compare.parse_error", msg=str(exc)))
            return
        self._file_path = path
        self.lbl_path.setText(os.path.basename(path))

        self.cmb_sheet.blockSignals(True)
        self.cmb_sheet.clear()
        for name in self._sheets:
            self.cmb_sheet.addItem(name)
        self.cmb_sheet.blockSignals(False)
        self.cmb_sheet.setCurrentIndex(0)
        self._on_sheet_changed(0)

    def _on_sheet_changed(self, idx: int) -> None:
        if idx < 0 or not self._sheets:
            return
        name = self.cmb_sheet.currentText()
        if name not in self._sheets:
            return
        headers, rows = self._sheets[name]
        self.table.set_rows(headers, rows)

    def get_points(self, fallback_radius: float) -> list[ComparePoint]:
        return self.table.get_mapped_points(fallback_radius)

    def refresh_texts(self) -> None:
        self.btn_browse.setText(t("compare.excel_select"))
        self.lbl_sheet.setText(t("compare.excel_sheet") + ":")
        self.table.refresh_texts()


# ── Main panel ───────────────────────────────────────────────────────────────

class ComparePanel(QGroupBox):
    """Panel for comparing land cover across multiple points."""

    # Emits list of ComparePoint instances
    compare_requested = Signal(list)
    open_map_requested = Signal(list)
    bulk_download_requested = Signal(str)
    map_tiff_requested = Signal(str)

    def minimumSizeHint(self) -> QSize:  # noqa: N802 - Qt API
        """Allow the form to reflow inside notebook-sized viewports."""
        return QSize(520, 420)

    def __init__(self, parent: QWidget | None = None):
        super().__init__(t("compare.title"), parent)
        self._point_rows: list[PointRow] = []
        self._last_results: list[dict] | None = None
        self._setup_ui()
        self._add_manual_point()
        self._add_manual_point()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setSpacing(8)

        # ── Input mode selector ─────────────────────────────────────────
        mode_row = QHBoxLayout()
        self.lbl_mode = QLabel(t("compare.input_mode") + ":")
        self.lbl_mode.setStyleSheet("font-weight: bold;")
        mode_row.addWidget(self.lbl_mode)

        self.rb_manual = QRadioButton(t("compare.mode_manual"))
        self.rb_paste = QRadioButton(t("compare.mode_paste"))
        self.rb_excel = QRadioButton(t("compare.mode_excel"))
        self.rb_manual.setChecked(True)

        self._mode_group = QButtonGroup(self)
        self._mode_group.addButton(self.rb_manual, 0)
        self._mode_group.addButton(self.rb_paste, 1)
        self._mode_group.addButton(self.rb_excel, 2)
        self._mode_group.idClicked.connect(self._on_mode_changed)

        mode_row.addWidget(self.rb_manual)
        mode_row.addWidget(self.rb_paste)
        mode_row.addWidget(self.rb_excel)
        mode_row.addStretch()
        root.addLayout(mode_row)

        # ── Stacked input area ─────────────────────────────────────────
        self._stack = QStackedWidget()
        self._stack.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred
        )

        # Manual page
        self._manual_page = QWidget()
        manual_layout = QVBoxLayout(self._manual_page)
        manual_layout.setContentsMargins(0, 0, 0, 0)
        self._points_container = QVBoxLayout()
        manual_layout.addLayout(self._points_container)
        self._btn_add_manual = QPushButton(t("compare.add"))
        self._btn_add_manual.clicked.connect(self._add_manual_point)
        add_row = QHBoxLayout()
        add_row.addWidget(self._btn_add_manual)
        add_row.addStretch()
        manual_layout.addLayout(add_row)
        self._stack.addWidget(self._manual_page)

        # Paste page
        self._paste_page = _PasteInput()
        self._stack.addWidget(self._paste_page)

        # Excel page
        self._excel_page = _ExcelInput()
        self._stack.addWidget(self._excel_page)

        root.addWidget(self._stack)

        # ── Unified radius option ──────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: #ddd;")
        root.addWidget(sep)

        radius_row = QHBoxLayout()
        self.chk_unified = QCheckBox(t("compare.unified_radius"))
        self.chk_unified.setChecked(True)
        self.chk_unified.toggled.connect(self._on_unified_toggled)
        radius_row.addWidget(self.chk_unified)

        self.lbl_unified = QLabel(t("compare.unified_radius_value") + ":")
        self.unified_radius = RadiusInput(default=5000.0)
        self.unified_radius.setMinimumWidth(120)
        radius_row.addWidget(self.lbl_unified)
        radius_row.addWidget(self.unified_radius)
        radius_row.addStretch()
        root.addLayout(radius_row)

        actions = QHBoxLayout()
        self._btn_open_map = QPushButton(t("compare_extra.show_map_btn"))
        self._btn_open_map.clicked.connect(self._on_open_map)
        self._btn_bulk_dl = QPushButton(t("compare_extra.bulk_download_btn"))
        self._btn_bulk_dl.clicked.connect(self._on_bulk_download)
        self._btn_map_tiff = QPushButton(t("compare_extra.map_tiff_btn"))
        self._btn_map_tiff.clicked.connect(self._on_map_tiff)
        actions.addWidget(self._btn_open_map)
        actions.addWidget(self._btn_bulk_dl)
        actions.addWidget(self._btn_map_tiff)
        actions.addStretch()
        root.addLayout(actions)

        chart_row = QHBoxLayout()
        self._lbl_gradient = QLabel("<b>" + t("compare_extra.gradient_title") + "</b>")
        self._lbl_grad_metric = QLabel(t("compare_extra.gradient_metric") + ":")
        self._cmb_grad_metric = QComboBox()
        self._cmb_grad_metric.addItem("ISA (%)", "isa")
        self._cmb_grad_metric.addItem("SHDI", "shdi")
        self._cmb_grad_metric.addItem("SIDI", "sidi")
        self._cmb_grad_metric.addItem("LPI", "lpi")
        self._cmb_grad_metric.addItem("Patches", "patches")
        self._cmb_grad_metric.currentIndexChanged.connect(self._redraw_gradient)
        chart_row.addWidget(self._lbl_gradient)
        chart_row.addSpacing(12)
        chart_row.addWidget(self._lbl_grad_metric)
        chart_row.addWidget(self._cmb_grad_metric)
        chart_row.addStretch()
        root.addLayout(chart_row)

        self._gradient_canvas = None
        if _MPL_OK:
            self._gradient_fig = Figure(figsize=(6, 2.6), tight_layout=True)
            self._gradient_canvas = _FigCanvas(self._gradient_fig)
            self._gradient_canvas.setMinimumHeight(180)
            self._gradient_canvas.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred,
            )
            root.addWidget(self._gradient_canvas)

        # ── Compare button ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._btn_compare = QPushButton(t("compare.compare_btn"))
        self._btn_compare.setStyleSheet(
            "QPushButton { background: #8e44ad; color: white; padding: 8px 28px; "
            "border-radius: 4px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background: #7d3c98; }"
        )
        self._btn_compare.clicked.connect(self._on_compare)
        btn_row.addWidget(self._btn_compare)
        root.addLayout(btn_row)

        # ── Results table ──────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        root.addWidget(self._table, 1)

        self._on_unified_toggled(True)

    # ── Mode handling ──────────────────────────────────────────────────

    def _on_mode_changed(self, mode_id: int) -> None:
        self._stack.setCurrentIndex(mode_id)

    def _on_unified_toggled(self, checked: bool) -> None:
        self.unified_radius.setEnabled(checked)
        self.lbl_unified.setEnabled(checked)
        # In manual mode, per-point radius inputs are shown when unified is OFF
        for row in self._point_rows:
            row.radius_input.setVisible(not checked)
            row.lbl_r.setVisible(not checked)

    # ── Manual rows ────────────────────────────────────────────────────

    def _add_manual_point(self) -> None:
        idx = len(self._point_rows)
        row = PointRow(idx, self)
        row.removed.connect(self._remove_manual_point)
        # Sync radius visibility with unified-radius state
        checked = self.chk_unified.isChecked() if hasattr(self, "chk_unified") else True
        row.radius_input.setVisible(not checked)
        row.lbl_r.setVisible(not checked)
        self._point_rows.append(row)
        self._points_container.addWidget(row)

    def _remove_manual_point(self, row: PointRow) -> None:
        if len(self._point_rows) <= 1:
            return
        self._points_container.removeWidget(row)
        row.deleteLater()
        self._point_rows.remove(row)
        # Re-index remaining rows for placeholder text
        for i, r in enumerate(self._point_rows):
            r.index = i
            r.refresh_texts()

    @property
    def gradient_metric(self) -> str:
        return self._cmb_grad_metric.currentData() or "isa"

    def _on_open_map(self) -> None:
        try:
            points = self._collect_points()
        except ValueError:
            return
        if points:
            self.open_map_requested.emit(points)

    def _on_bulk_download(self) -> None:
        if not self._last_results:
            QMessageBox.warning(self, "", t("compare_extra.bulk_download_no_data"))
            return
        path = QFileDialog.getExistingDirectory(
            self, t("compare_extra.bulk_download_dir")
        )
        if path:
            self.bulk_download_requested.emit(path)

    def _on_map_tiff(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, t("compare_extra.map_tiff_btn"), "luma_compare_map.tif",
            "TIFF (*.tif *.tiff)",
        )
        if path:
            self.map_tiff_requested.emit(path)

    def _redraw_gradient(self) -> None:
        if not _MPL_OK or self._gradient_canvas is None or not self._last_results:
            return
        accessors = {
            "isa": lambda metrics: metrics.isa_index,
            "shdi": lambda metrics: metrics.shannon_diversity,
            "sidi": lambda metrics: metrics.simpson_diversity,
            "lpi": lambda metrics: metrics.largest_patch_index,
            "patches": lambda metrics: metrics.total_patches,
        }
        values = [
            accessors[self.gradient_metric](result["landscape_metrics"])
            for result in self._last_results
        ]
        labels = [result["point_label"] for result in self._last_results]
        self._gradient_fig.clear()
        axis = self._gradient_fig.add_subplot(111)
        import numpy as np
        x = np.arange(len(labels))
        minimum, maximum = min(values), max(values)
        span = maximum - minimum or 1.0
        colors = [
            (1 - (value - minimum) / span, 0.4, (value - minimum) / span)
            for value in values
        ]
        axis.bar(x, values, color=colors, edgecolor="#222")
        axis.plot(x, values, color="#222", marker="o", linewidth=1)
        axis.set_xticks(x)
        axis.set_xticklabels(labels, rotation=30, ha="right", fontsize=8)
        axis.set_ylabel(self._cmb_grad_metric.currentText())
        axis.grid(axis="y", alpha=0.25)
        self._gradient_canvas.draw_idle()

    # ── Compare ────────────────────────────────────────────────────────

    def _on_compare(self) -> None:
        try:
            points = self._collect_points()
        except ValueError as exc:
            QMessageBox.warning(self, "Error", t("compare.parse_error", msg=str(exc)))
            return

        if not points:
            QMessageBox.warning(self, "Error", t("compare.no_points"))
            return

        self.compare_requested.emit(points)

    def _collect_points(self) -> list[ComparePoint]:
        mode = self._stack.currentIndex()
        unified = self.chk_unified.isChecked()
        fallback_r = self.unified_radius.value()

        if mode == 0:  # manual
            out: list[ComparePoint] = []
            for row in self._point_rows:
                p = row.get_point()
                if unified:
                    p.radius = fallback_r
                out.append(p)
            return out
        elif mode == 1:  # paste
            pts = self._paste_page.get_points(fallback_r)
            if unified:
                for p in pts:
                    p.radius = fallback_r
            return pts
        else:  # excel
            pts = self._excel_page.get_points(fallback_r)
            if unified:
                for p in pts:
                    p.radius = fallback_r
            return pts

    # ── Results rendering ──────────────────────────────────────────────

    def update_results(self, results: list[dict]) -> None:
        """Display comparison table.

        Each dict has: 'point_label', 'class_stats', 'landscape_metrics'.
        """
        if not results:
            return
        self._last_results = results

        all_classes = []
        seen = set()
        for r in results:
            for cs in r["class_stats"]:
                if cs.class_name not in seen:
                    all_classes.append(cs.class_name)
                    seen.add(cs.class_name)

        metric_defs = [
            ("compare.metric_shdi",          lambda m: f"{m.shannon_diversity:.3f}"),
            ("compare.metric_isa",           lambda m: f"{m.isa_index:.1f}%"),
            ("compare.metric_sidi",          lambda m: f"{m.simpson_diversity:.3f}"),
            ("compare.metric_evenness",      lambda m: f"{m.evenness:.3f}"),
            ("compare.metric_patches",       lambda m: f"{m.total_patches}"),
            ("compare.metric_patch_density", lambda m: f"{m.patch_density:.1f}"),
            ("compare.metric_lpi",           lambda m: f"{m.largest_patch_index:.1f}"),
            ("compare.metric_aggregation",   lambda m: f"{m.aggregation_index:.1f}"),
            ("compare.metric_contagion",     lambda m: f"{m.contagion:.1f}" if m.contagion is not None else "—"),
            ("compare.metric_shape",         lambda m: f"{m.mean_shape_index:.3f}"),
            ("compare.metric_area_max",      lambda m: f"{m.largest_patch_area_m2 / 10_000:,.2f}"),
            ("compare.metric_area_min",      lambda m: f"{m.smallest_patch_area_m2 / 10_000:,.2f}"),
            ("compare.metric_area_mean",     lambda m: f"{m.mean_patch_area_m2 / 10_000:,.2f}"),
        ]

        # rows = points, cols = classes + metrics
        n_rows = len(results)
        n_cols = len(all_classes) + len(metric_defs) + 1
        self._table.setRowCount(n_rows)
        self._table.setColumnCount(n_cols)

        col_headers = [""] + all_classes + [t(key) for key, _ in metric_defs]
        self._table.setHorizontalHeaderLabels(col_headers)

        for row_idx, r in enumerate(results):
            self._table.setItem(row_idx, 0, QTableWidgetItem(r["point_label"]))
            for col_offset, cls_name in enumerate(all_classes):
                pct = 0.0
                for cs in r["class_stats"]:
                    if cs.class_name == cls_name:
                        pct = cs.percentage
                        break
                self._table.setItem(row_idx, col_offset + 1, QTableWidgetItem(f"{pct:.1f}%"))
            for col_offset, (_, accessor) in enumerate(metric_defs):
                col_idx = len(all_classes) + 1 + col_offset
                self._table.setItem(row_idx, col_idx, QTableWidgetItem(accessor(r["landscape_metrics"])))

        self._table.resizeColumnsToContents()
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._redraw_gradient()

    # ── i18n refresh ───────────────────────────────────────────────────

    def refresh_texts(self) -> None:
        self.setTitle(t("compare.title"))
        self.lbl_mode.setText(t("compare.input_mode") + ":")
        self.rb_manual.setText(t("compare.mode_manual"))
        self.rb_paste.setText(t("compare.mode_paste"))
        self.rb_excel.setText(t("compare.mode_excel"))
        self._btn_add_manual.setText(t("compare.add"))
        self._btn_compare.setText(t("compare.compare_btn"))
        self.chk_unified.setText(t("compare.unified_radius"))
        self.lbl_unified.setText(t("compare.unified_radius_value") + ":")
        self._btn_open_map.setText(t("compare_extra.show_map_btn"))
        self._btn_bulk_dl.setText(t("compare_extra.bulk_download_btn"))
        self._btn_map_tiff.setText(t("compare_extra.map_tiff_btn"))
        self._lbl_gradient.setText("<b>" + t("compare_extra.gradient_title") + "</b>")
        self._lbl_grad_metric.setText(t("compare_extra.gradient_metric") + ":")
        for row in self._point_rows:
            row.refresh_texts()
        self._paste_page.refresh_texts()
        self._excel_page.refresh_texts()
        if self._last_results:
            self.update_results(self._last_results)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    """Parse a float accepting both '.' and ',' as decimal separators."""
    if val is None:
        raise ValueError("empty")
    s = str(val).strip()
    if not s:
        raise ValueError("empty")
    # Common European/Brazilian format: "1.234,56" -> remove thousands
    if "," in s and "." in s:
        # Assume '.' is thousand separator
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    value = float(s)
    if not math.isfinite(value):
        raise ValueError("non-finite")
    return value


def _parse_delimited(raw: str) -> tuple[list[str], list[list[str]]]:
    """Parse a pasted table. Auto-detects delimiter (tab, comma, semicolon)."""
    # Sniff delimiter on the first non-empty line
    lines = [ln for ln in raw.splitlines() if ln.strip()]
    if not lines:
        return [], []
    sample = "\n".join(lines[:5])
    delim = "\t"
    try:
        sniffed = csv.Sniffer().sniff(sample, delimiters="\t,;|")
        delim = sniffed.delimiter
    except csv.Error:
        # Fallback: pick whichever appears most on the first line
        counts = {d: lines[0].count(d) for d in ["\t", ";", ",", "|"]}
        delim = max(counts, key=counts.get) if max(counts.values()) > 0 else ","

    reader = csv.reader(io.StringIO(raw), delimiter=delim)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return [], []

    headers = [h.strip() for h in rows[0]]
    data_rows = [row for row in rows[1:]]
    return headers, data_rows


def _read_spreadsheet(path: str) -> dict[str, tuple[list[str], list[list[str]]]]:
    """Read an .xlsx/.xls/.csv file into {sheet_name: (headers, rows)}."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            raw = f.read()
        headers, rows = _parse_delimited(raw)
        return {"Sheet1": (headers, rows)}

    # Excel — prefer openpyxl for .xlsx
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required to read Excel files. Install with: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, tuple[list[str], list[list[str]]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            sheets[name] = ([], [])
            continue
        headers = [str(h) if h is not None else "" for h in first]
        data_rows: list[list[str]] = []
        for row in rows_iter:
            if row is None:
                continue
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            data_rows.append(["" if c is None else str(c) for c in row])
        sheets[name] = (headers, data_rows)
    wb.close()
    return sheets
